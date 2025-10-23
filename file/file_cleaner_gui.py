#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Cleaner & Standardizer - versión con log persistente (Excel/CSV) y vista previa separada.
Guarda/actualiza un archivo de log existente o lo crea si no existe.
"""

import os
import re
import unicodedata
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Importar pandas para manejo Excel/CSV; si no está, usa CSV puro.
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except Exception:
    PANDAS_AVAILABLE = False

# ---------------------------
# Utilidades
# ---------------------------
def ascii_clean(name: str, lower: bool = True) -> str:
    """sanitiza un nombre a ASCII, reemplazando caracteres no permitidos por guiones bajos.

    Args:
        name (str): _nombre a limpiar_
        lower (bool, optional): _si convertir a minusculas_. Defaults to True.

    Returns:
        str: _nombre limpio_
    """
    if not name:
        return ''

    # mapeo explícito de caracteres problemáticos a su equivalente textual
    CHAR_MAP = {
        '¢': 'o',
        '©': 'c',
        '®': 'r',
        '–': '-',  # guion largo a guion medio
        '—': '-',  # em dash
        '’': "'",  # apóstrofe tipográfico a simple
        '‘': "'",
        '“': '"',
        '”': '"',
        'ß': 'ss'
    }
        
    s = name
    for src, tgt in CHAR_MAP.items():
        s = s.replace(src, tgt)

    n = unicodedata.normalize('NFKD', s)
    n = n.encode('ASCII', 'ignore').decode('ASCII')
    # Reemplazar caracteres que no sean letras/números/._- por _
    n = re.sub(r'[^\w\.\-]', '_', n)
    n = re.sub(r'_+', '_', n)
    n = n.strip(' _')
    return n.lower() if lower else n

def process_name_for_standard(name: str, move_leading_number: bool = True, number_sep: str = '_') -> tuple[str, str|None]:
    """
    Normaliza la parte NAME para el estándar:
        - convierte acentos a ascii (ascii_clean sin forzar lower)
        - reemplaza '-' por '_' y todos los '.' (puntos) internos por '_'
        - colapsa '_' repetidos
        - opcional: si name empieza con una secuencia numérica (ej: '123', '24 y 25', '8-21', '24,25'),
        extrae esa secuencia y devuelve (name_sin_numero, normalized_number) donde normalized_number
        es la secuencia con separadores convertidos a '_', por ejemplo '24 y 25' -> '24_25'.
    Devuelve (processed_name, moved_number_or_None)
    """
    if not name:
        return '', None

    # 1) normalizar acentos sin forzar lowercase
    n = ascii_clean(name, lower=False)

    moved_number = None
    if move_leading_number:
        # Regex para capturar secuencias de números al inicio separados por -,–,.,, y, and, &
        # Ej: "24 y 25 R.AP..." -> group1 = "24 y 25", rest = "R.AP..."
        m = re.match(r'^\s*([0-9]+(?:\s*(?:[\-–\.,]|y|and|&)\s*[0-9]+)*)\s*[-_\.\s:]*(.+)$',
                    n, flags=re.IGNORECASE)
        if m:
            raw_num_seq = m.group(1)
            rest = m.group(2)
            # normalizar separadores a '_' y quitar espacios sobrantes
            normalized = re.sub(r'[\s]*?(?:[\-–\.,]|y|and|&)[\s]*?', '_', raw_num_seq, flags=re.IGNORECASE)
            normalized = re.sub(r'_+', '_', normalized).strip('_ ')
            moved_number = normalized if normalized else None
            n = rest

    # 2) reemplazar guion medio por underscore
    n = n.replace('-', '_')

    # 3) reemplazar todos los puntos por underscore (estos puntos están en el 'name' ya separado de la extensión)
    n = n.replace('.', '_')

    # 4) colapsar múltiples underscores y limpiar bordes
    n = re.sub(r'_+', '_', n).strip('_ ')

    return n, moved_number

def capitalize_first_word_keep_underscores(name: str) -> str:
    """
    Pone todo en minúsculas y capitaliza sólo la primera letra de la PRIMERA palabra.
    Palabras separadas por '_' se mantienen así.
    Ej: 'HOJA_DE_VIDA' -> 'Hoja_de_vida'
    """
    if not name:
        return name
    lower = name.lower()
    parts = lower.split('_')
    if parts:
        parts[0] = parts[0].capitalize()
    return '_'.join(parts)

def safe_unique_path(directory: str, filename: str) -> str:
    """Genera una ruta unica en el directorio dado, añadiendo sufijos numericos si es necesario.

    Args:
        directory (str): _directorio donde comprobar_
        filename (str): _nombre de archivo deseado_

    Returns:
        str: _ruta única disponible_
    """
    base, ext = os.path.splitext(filename)
    candidate = filename
    i = 1
    while os.path.exists(os.path.join(directory, candidate)):
        candidate = f"{base}_{i}{ext}"
        i += 1
    return os.path.join(directory, candidate)

def is_temp_file(filename: str, temp_patterns=None) -> bool:
    """Determina si un archivo es temporal o erróneo segun patrones predefinidos.
    Args:
        filename (str): _nombre de archivo_
        temp_patterns (list, optional): _lista de patrones regex para identificar archivos temporales_. Defaults to None.
    Returns:
        bool: _True si es archivo temporal, False en caso contrario_
    """
    if temp_patterns is None:
        temp_patterns = [
            r'^\~', r'\.tmp$', r'\.bak$', r'\.old$', r'^\._', r'\.swp$',
            r'^Thumbs\.db$', r'^\.DS_Store$', r'^\~\$'
        ]
    for pat in temp_patterns:
        if re.search(pat, filename, flags=re.IGNORECASE):
            return True
    return False

def parse_mapping_text(text: str) -> dict:
    """Mapea texto multilinea en un diccionario clave:valor. Cada linea debe tener formato 'clave:valor'.
    
    Args:
        text (str): _texto a parsear_
    Returns:
        dict: _diccionario resultante_
    """
    mapping = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if ':' in line:
            key, val = line.split(':', 1)
            mapping[key.strip().lower()] = val.strip()
    return mapping

def find_area_abbr_in_path(path: str, mapping: dict) -> str:
    """Busca en la ruta las carpetas que coincidan con las claves del mapping y devuelve la abreviación asociada.
    Args:
        path (str): _ruta a analizar_
        mapping (dict): _diccionario clave:abreviación_
    Returns:
        str: _abreviación encontrada o cadena vacía_
    """
    parts = [p.lower() for p in path.replace('\\','/').split('/') if p]
    for part in parts[::-1]:
        if part in mapping:
            return mapping[part]
    return ''

def build_standard_name(original_filename: str, folder_path: str,
                        pattern: str, prefix_choice: str, area_abbr_map: dict) -> tuple[str, str]:
    """
    Construye un nombre de archivo estándar según el patrón y los parámetros dados.
    Ahora:
        - process_name_for_standard normaliza NAME (reemplaza '-', '.' internos por '_'),
        - si se detecta número al inicio, lo mueve y lo agrega como sufijo del NAME (antes de la extensión)
    """
    name, ext = os.path.splitext(original_filename)
    ext = ext.lstrip('.').lower()

    # 1) intentar extraer prefijo+area si vienen al inicio del filename (ej "R.AP Paola..." o "R AP Paola...")
    detected_prefix = ''
    detected_area = ''
    rest_name = name

    m_pref_area = re.match(r'^\s*([A-Za-z0-9]{1,6})[.\s]+([A-Za-z0-9]{1,6})[.\s]+(.+)$', name)
    if m_pref_area:
        # ej: 'R.AP Paola...' -> pref='R', area='AP', rest='Paola...'
        detected_prefix = m_pref_area.group(1)
        detected_area = m_pref_area.group(2)
        rest_name = m_pref_area.group(3)
    else:
        # también soportar 'R.AP.' (puntos) sin espacio: 'R.AP.Paola...'
        m2 = re.match(r'^\s*([A-Za-z0-9]{1,6})\.([A-Za-z0-9]{1,6})\.(.+)$', name)
        if m2:
            detected_prefix = m2.group(1)
            detected_area = m2.group(2)
            rest_name = m2.group(3)

    # 2) detectar sufijo numérico justo antes de la extensión (ej '...Nombre.02' -> suffix '02')
    moved_number = None
    m_suffix_num = re.match(r'^(.*?)[\.\s_-]*([0-9]+)$', rest_name)
    if m_suffix_num:
        rest_name = m_suffix_num.group(1)
        moved_number = m_suffix_num.group(2)

    # 3) normalizar NAME con la función procesadora (extrae también números compuestos si están al inicio)
    processed_name, moved_from_start = process_name_for_standard(rest_name, move_leading_number=True)
    processed_name = capitalize_first_word_keep_underscores(processed_name)
    # si había número movido del inicio, añadirlo al final (como ya hacíamos)
    number_parts = []
    if moved_from_start:
        number_parts.append(moved_from_start)
    if moved_number:
        number_parts.append(moved_number)
    if number_parts:
        # unir con _
        suffix = "_".join(number_parts)
        if processed_name:
            processed_name = f"{processed_name}{'_' if not processed_name.endswith('_') else ''}{suffix}"
        else:
            processed_name = suffix

    # escoger prefix/area: si detectamos en el filename, preferirlos; sino usar prefix_choice y area_abbr_map
    prefix_final = detected_prefix if detected_prefix else (prefix_choice or '')
    # obtener area por mapeo (si detectada es texto, intentar mapear a abreviación)
    if detected_area:
        # si detected_area es ya la abreviatura (p ej 'AP' o 'CP') la tomamos; si no, intentar mapear lowercase key
        a = detected_area.strip()
        # buscar en mapping keys por coincidencia simple (case-insensitive)
        found_abbr = ''
        for k, v in area_abbr_map.items():
            if a.lower() == k.lower() or a.lower() == v.lower():
                found_abbr = v
                break
        if found_abbr:
            area_final = found_abbr
        else:
            area_final = find_area_abbr_in_path(folder_path, area_abbr_map) or a
    else:
        area_final = find_area_abbr_in_path(folder_path, area_abbr_map) or ''

    # construir reemplazos
    parent = os.path.basename(folder_path) or ''
    parent_letter = parent[0].upper() if parent else ''

    replacements = {
        '{PREFIX}': prefix_final,
        '{AREA}': area_final,
        '{NAME}': processed_name,
        '{EXT}': ext,
        '{PARENT_LETTER}': parent_letter
    }
    newname = pattern
    for k, v in replacements.items():
        newname = newname.replace(k, v)

    newname = re.sub(r'\.+', '.', newname)
    newname = re.sub(r'_+', '_', newname)
    newname = newname.strip('. _')

    if '{EXT}' not in pattern and ext:
        newname = f"{newname}.{ext}"

    note = f"area:{area_final} parent:{parent_letter}"
    return newname, note

# ---------------------------
# Log persistente: append a XLSX/CSV
# ---------------------------
def append_log_entries_to_file(entries: list, path: str):
    """Agrega entradas de log a un archivo Excel (.xlsx) o CSV (.csv). Si el archivo no existe, lo crea.
    Args:
        entries (list): _lista de entradas de log (dicts)_
        path (str): _ruta del archivo de log_
    """
    if not entries:
        return
    # Normalizar keys y orden
    keys = ['timestamp', 'original_path', 'action', 'new_path', 'note']

    def relativize(p: str) -> str:
        if not p:
            return ''
        # normalizar separadores
        pp = p.replace('\\', '/')
        # buscar 'trunk' (caso-insensible)
        idx = pp.lower().find('/trunk/')
        if idx >= 0:
            return pp[idx+1:]  # devuelve inicio en 'trunk/...'
        idx2 = pp.lower().find('trunk/')
        if idx2 >= 0:
            return pp[idx2:]   # por si no tiene prefijo slash
        # si no encuentra 'trunk', devolver path relativo al drive (sin cambios)
        return pp
        
    # crear copia transformada de entries con rutas relativas
    entries_rel = []
    for e in entries:
        e2 = dict(e)  # copia
        e2['original_path'] = relativize(e.get('original_path', ''))
        e2['new_path'] = relativize(e.get('new_path', ''))
        entries_rel.append(e2)
    
    ext = os.path.splitext(path)[1].lower()
    # Try pandas/openpyxl as before
    pandas_available_local = False
    try:
        import pandas as pd
        pandas_available_local = True
    except Exception:
        pandas_available_local = False

    if ext == '.xlsx' and pandas_available_local:
        import pandas as pd
        df_new = pd.DataFrame(entries_rel).reindex(columns=keys)
        if os.path.exists(path):
            try:
                df_old = pd.read_excel(path)
                df_combined = pd.concat([df_old, df_new], ignore_index=True)
                df_combined.to_excel(path, index=False)
            except Exception:
                df_new.to_excel(path, index=False)
        else:
            df_new.to_excel(path, index=False)
        return

    if ext == '.xlsx' and not pandas_available_local:
        try:
            from openpyxl import Workbook, load_workbook
            if os.path.exists(path):
                wb = load_workbook(path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(keys)
            for e in entries_rel:
                ws.append([e.get(k, '') for k in keys])
            wb.save(path)
            return
        except Exception:
            pass

    # fallback CSV
    csv_path = path if ext == '.csv' else os.path.splitext(path)[0] + '.csv'
    write_header = not os.path.exists(csv_path)
    import csv
    with open(csv_path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        if write_header:
            writer.writeheader()
        for e in entries_rel:
            writer.writerow({k: e.get(k, '') for k in keys})

def pattern_to_regex(pattern: str):
    """Convierte un pattern con placeholders en una regex con named groups.
    Los placeholders son: {PREFIX}, {AREA}, {NAME}, {EXT}, {PARENT_LETTER}
    y se convierten en grupos regex seguros.
    Args:
        pattern (str): _patrón con placeholders_
    Returns:
        re.Pattern: _objeto regex compilado_    
    """
    # escapamos los caracteres especiales del pattern excepto los placeholders
    # reemplazamos placeholders por subregexes seguras
    regex = re.escape(pattern)

    replacements = {
        re.escape('{PREFIX}'): r'(?P<PREFIX>[^.]+)',
        re.escape('{AREA}'): r'(?P<AREA>[^.]+)',
        re.escape('{NAME}'): r'(?P<NAME>.+)',
        re.escape('{EXT}'): r'(?P<EXT>[^.]+)',
        re.escape('{PARENT_LETTER}'): r'(?P<PARENT_LETTER>.)'
    }
    for ph_esc, repl in replacements.items():
        regex = regex.replace(ph_esc, repl)
    regex = '^' + regex + '$'
    return re.compile(regex)

def shorten_path(path: str, show_parts: int = 3) -> str:
    """
    Devuelve una versión 'resumida' de la ruta:
    muestra las últimas `show_parts` partes separadas por '/' (o '\' en Windows),
    pero presentadas como: parent2 / parent1 / filename
    Si la ruta tiene menos partes, muestra la ruta completa.
    Ej: C:/.../Tesoreria/Registros/R.TR...xlsx -> Tesoreria / Registros / R.TR...xlsx
    """
    if not path:
        return ''
    sep = '/' if '/' in path else '\\'
    parts = [p for p in path.replace('\\','/').split('/') if p]
    if not parts:
        return path
    # tomar las últimas show_parts partes
    last = parts[-show_parts:]
    return " / ".join(last)

# ---------------------------
# App GUI
# ---------------------------
class FileCleanerApp:
    def __init__(self, root):
        import sys  # <- asegurarse que sys esté importado (también puedes colocarlo arriba del archivo)
        self.root = root
        root.title("File Cleaner & Standardizer - Con log persistente")
        self.session_preview = []  # preview entries
        self.session_applied = []  # applied entries (para escribir luego)

        # --- MAIN scrollable canvas (envuelve toda la UI) ---
        root.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)

        self.main_canvas = tk.Canvas(root)
        vscroll_main = ttk.Scrollbar(root, orient='vertical', command=self.main_canvas.yview)
        self.main_canvas.configure(yscrollcommand=vscroll_main.set)

        self.main_canvas.grid(row=0, column=0, sticky='nsew')
        vscroll_main.grid(row=0, column=1, sticky='ns')

        # frame interior que contendrá toda la UI
        self.main_frame = ttk.Frame(self.main_canvas)
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.main_frame, anchor='nw')

        # ajustar scrollregion cuando cambie el tamaño del contenido
        def _on_frame_config(e):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        self.main_frame.bind("<Configure>", _on_frame_config)

        # mantener el ancho del frame interior igual al ancho del canvas (pero solo si cambia)
        self._last_canvas_width = None
        def _on_canvas_config(e):
            # e.width puede ser 1 en algunos momentos; mejor consultar winfo_width
            w = self.main_canvas.winfo_width()
            if w <= 1:
                return
            if self._last_canvas_width == w:
                return
            try:
                self.main_canvas.itemconfigure(self.canvas_window, width=w)
            except Exception:
                pass
            self._last_canvas_width = w
        # enlazamos al configure del propio canvas (no al root) para evitar demasiadas llamadas y "parpadeos"
        self.main_canvas.bind("<Configure>", _on_canvas_config)

        # ----------------------------------------------------
        # Ahora construimos la UI dentro de self.main_frame
        # ----------------------------------------------------
        frm_top = ttk.Frame(self.main_frame, padding=8)
        frm_top.grid(row=0, column=0, columnspan=2, sticky='ew')
        ttk.Label(frm_top, text="Carpeta raíz:").grid(row=0, column=0, sticky='w')
        self.folder_var = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.folder_var, width=60).grid(row=0, column=1, padx=6, sticky='ew')
        ttk.Button(frm_top, text="Seleccionar...", command=self.select_folder).grid(row=0, column=2)

        # SUBDIRS panel
        frm_subdirs = ttk.LabelFrame(self.main_frame, text="Subcarpetas (seleccionar donde aplicar)", padding=8)
        frm_subdirs.grid(row=1, column=1, sticky='nsew', padx=8, pady=6)
        frm_subdirs.columnconfigure(0, weight=1)
        self.subdirs_canvas = tk.Canvas(frm_subdirs, width=320, height=200)
        self.subdirs_inner = ttk.Frame(self.subdirs_canvas)
        vsb = ttk.Scrollbar(frm_subdirs, orient="vertical", command=self.subdirs_canvas.yview)
        self.subdirs_canvas.configure(yscrollcommand=vsb.set)
        vsb.grid(row=0, column=1, sticky='ns')
        self.subdirs_canvas.grid(row=0, column=0, sticky='nsew')
        self.subdirs_canvas_window = self.subdirs_canvas.create_window((0,0), window=self.subdirs_inner, anchor='nw')
        self.subdirs_inner.bind("<Configure>", lambda e: self.subdirs_canvas.configure(scrollregion=self.subdirs_canvas.bbox("all")))
        self.folder_checks = {}

        btns_frame = ttk.Frame(frm_subdirs)
        btns_frame.grid(row=1, column=0, columnspan=2, pady=(6,0), sticky='ew')
        ttk.Button(btns_frame, text="Seleccionar todo", command=self.select_all_subdirs).grid(row=0, column=0, padx=2)
        ttk.Button(btns_frame, text="Deseleccionar todo", command=lambda: self.select_all_subdirs(select=False)).grid(row=0, column=1, padx=2)
        ttk.Button(btns_frame, text="Invertir", command=self.invert_subdirs).grid(row=0, column=2, padx=2)

        # OPTIONS (izquierda)
        frm_opts = ttk.LabelFrame(self.main_frame, text="Opciones", padding=8)
        frm_opts.grid(row=1, column=0, sticky='ew', padx=8, pady=6)
        self.delete_vars = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_opts, text="Eliminar archivos temporales/erróneos", variable=self.delete_vars).grid(row=0, column=0, sticky='w')
        self.clean_names_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_opts, text="Limpiar nombres (acentos, caracteres especiales)", variable=self.clean_names_var).grid(row=1, column=0, sticky='w')
        self.apply_standard_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_opts, text="Aplicar/validar estándar", variable=self.apply_standard_var).grid(row=2, column=0, sticky='w')

        # mantener proporciones de columnas dentro del main_frame
        self.main_frame.columnconfigure(0, weight=3)
        self.main_frame.columnconfigure(1, weight=1)

        # STANDARD pattern
        frm_std = ttk.LabelFrame(self.main_frame, text="Estándar (pattern)", padding=8)
        frm_std.grid(row=2, column=0, sticky='ew', padx=8, pady=6)
        ttk.Label(frm_std, text="Pattern (placeholders: {PREFIX},{AREA},{NAME},{EXT},{PARENT_LETTER}):").grid(row=0, column=0, sticky='w')
        self.pattern_var = tk.StringVar(value="{PARENT_LETTER}.{AREA}.{NAME}.{EXT}")
        ttk.Entry(frm_std, textvariable=self.pattern_var, width=70).grid(row=1, column=0, sticky='w', pady=4)
        ttk.Label(frm_std, text="Prefijo (ej: F, P, G, I, IND, RUT):").grid(row=2, column=0, sticky='w')
        self.prefix_var = tk.StringVar(value="P")
        ttk.Entry(frm_std, textvariable=self.prefix_var, width=10).grid(row=3, column=0, sticky='w', pady=4)

        # MAPPING
        frm_map = ttk.LabelFrame(self.main_frame, text="Mapeo de áreas -> abreviación (línea por línea 'NombreCarpeta:ABR')", padding=8)
        frm_map.grid(row=3, column=0, sticky='ew', padx=8, pady=6)
        self.map_text = scrolledtext.ScrolledText(frm_map, height=6, width=80)
        self.map_text.insert('1.0', "Gestión Humana:CP\nAdministración del personal:AP\nContratacion de Personal:CP")
        self.map_text.grid(row=0, column=0)

        # LOG path
        frm_log = ttk.LabelFrame(self.main_frame, text="Archivo de log (persistente)", padding=8)
        frm_log.grid(row=4, column=0, sticky='ew', padx=8, pady=6)
        ttk.Label(frm_log, text="Ruta archivo log (.xlsx o .csv):").grid(row=0, column=0, sticky='w')
        self.log_path_var = tk.StringVar()
        ttk.Entry(frm_log, textvariable=self.log_path_var, width=60).grid(row=1, column=0, sticky='w', padx=6)
        ttk.Button(frm_log, text="Seleccionar/Crear archivo log...", command=self.select_log_file).grid(row=1, column=1, padx=6)
        ttk.Label(frm_log, text="(Si no existe, se creará. Si no tienes pandas se usará CSV)").grid(row=2, column=0, sticky='w', pady=4)

        # BUTTONS
        frm_buttons = ttk.Frame(self.main_frame, padding=8)
        frm_buttons.grid(row=5, column=0, columnspan=2, sticky='ew')
        ttk.Button(frm_buttons, text="Generar vista previa", command=lambda: self.run_scan(mode='preview')).grid(row=0, column=0, padx=6)
        ttk.Button(frm_buttons, text="Aplicar cambios (escribirá en disco y actualizará log)", command=lambda: self.run_scan(mode='apply')).grid(row=0, column=1, padx=6)
        ttk.Button(frm_buttons, text="Exportar log sesión (temporal)", command=self.export_session_log).grid(row=0, column=2, padx=6)
        ttk.Button(frm_buttons, text="Limpiar panel", command=self.clear_panel).grid(row=0, column=3, padx=6)

        # RESULTADOS - ocupando ancho completo
        frm_result = ttk.LabelFrame(self.main_frame, text="Resultados / Acciones detectadas", padding=8)
        frm_result.grid(row=6, column=0, columnspan=2, sticky='nsew', padx=8, pady=6)
        frm_result.rowconfigure(0, weight=1)
        frm_result.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(frm_result, columns=('timestamp','original', 'action', 'new', 'note'),
                                show='headings', height=28)
        self.tree.grid(row=0, column=0, sticky='nsew', padx=(0,0), pady=(0,0))

        # barras
        vsb2 = ttk.Scrollbar(frm_result, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=vsb2.set)
        vsb2.grid(row=0, column=1, sticky='ns')
        hsb2 = ttk.Scrollbar(frm_result, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(xscroll=hsb2.set)
        hsb2.grid(row=1, column=0, columnspan=2, sticky='ew')

        # encabezados y columnas
        self.tree.heading('timestamp', text='Timestamp')
        self.tree.heading('original', text='Archivo original')
        self.tree.heading('action', text='Acción')
        self.tree.heading('new', text='Nuevo / Ruta')
        self.tree.heading('note', text='Nota')
        self.tree.column('timestamp', width=160, stretch=False)
        self.tree.column('original', width=700, stretch=True)
        self.tree.column('action', width=140, stretch=False)
        self.tree.column('new', width=480, stretch=True)
        self.tree.column('note', width=300, stretch=True)

        # ---------------------------
        # MOUSE WHEEL: comportamiento por plataforma
        # ---------------------------
        # objetivo del scroll: si el cursor está sobre tree => scroll tree,
        # si está sobre subdirs_canvas => scroll subdirs, sino scroll main_canvas.

        self._wheel_target = None

        def _enter_widget(e):
            # guardamos la referencia del widget donde está el cursor
            self._wheel_target = e.widget

        def _leave_widget(e):
            # solo limpiar si se sale del widget trackeado
            if self._wheel_target == e.widget:
                self._wheel_target = None

        # bind enter/leave para widgets relevantes
        widgets_for_wheel = [self.main_canvas, self.subdirs_canvas, self.subdirs_inner, self.tree]
        for w in widgets_for_wheel:
            try:
                w.bind("<Enter>", _enter_widget, add=True)
                w.bind("<Leave>", _leave_widget, add=True)
            except Exception:
                pass

        import platform
        is_linux = platform.system().lower().startswith('linux')

        def _mousewheel_handler(event):
            # decidir destino
            if self._wheel_target is not None:
                tgt = self._wheel_target
            else:
                tgt = self.main_canvas

            # eventos distintos según plataforma
            if is_linux:
                # en X11 los eventos son Button-4 (up) / Button-5 (down) y no usan event.delta
                return  # handled elsewhere by button bindings
            else:
                # Windows / macOS usan MouseWheel con event.delta
                if tgt == self.tree:
                    if sys.platform.startswith('win'):
                        self.tree.yview_scroll(-1 * (event.delta // 120), "units")
                    else:
                        # macOS suele usar event.delta con valores pequeños
                        self.tree.yview_scroll(-1 * int(event.delta), "units")
                elif tgt == self.subdirs_canvas or tgt == self.subdirs_inner:
                    if sys.platform.startswith('win'):
                        self.subdirs_canvas.yview_scroll(-1 * (event.delta // 120), "units")
                    else:
                        self.subdirs_canvas.yview_scroll(-1 * int(event.delta), "units")
                else:
                    if sys.platform.startswith('win'):
                        self.main_canvas.yview_scroll(-1 * (event.delta // 120), "units")
                    else:
                        self.main_canvas.yview_scroll(-1 * int(event.delta), "units")

        # para Windows / macOS enlazamos MouseWheel globalmente (pero actuará según self._wheel_target)
        root.bind_all("<MouseWheel>", _mousewheel_handler)

        # para Linux (X11) enlazamos Button-4 / Button-5 (rueda)
        def _linux_scroll_up(e):
            tgt = self._wheel_target or self.main_canvas
            if tgt == self.tree:
                self.tree.yview_scroll(-1, "units")
            elif tgt == self.subdirs_canvas or tgt == self.subdirs_inner:
                self.subdirs_canvas.yview_scroll(-1, "units")
            else:
                self.main_canvas.yview_scroll(-1, "units")
        def _linux_scroll_down(e):
            tgt = self._wheel_target or self.main_canvas
            if tgt == self.tree:
                self.tree.yview_scroll(1, "units")
            elif tgt == self.subdirs_canvas or tgt == self.subdirs_inner:
                self.subdirs_canvas.yview_scroll(1, "units")
            else:
                self.main_canvas.yview_scroll(1, "units")

        root.bind_all("<Button-4>", _linux_scroll_up)
        root.bind_all("<Button-5>", _linux_scroll_down)

    def select_folder(self):
        """Muestra diálogo para seleccionar carpeta raíz y poblar subcarpetas."""
        folder = filedialog.askdirectory()
        if folder:
            self.folder_var.set(folder)
            self.populate_folder_checkboxes(folder)

    def select_log_file(self):
        """Muestra diálogo para seleccionar o crear archivo de log (.xlsx o .csv)."""
        # mostrar save as para .xlsx o .csv
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files","*.*")])
        if path:
            self.log_path_var.set(path)

    def clear_panel(self):
        """Limpia el panel de resultados y las sesiones internas."""
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.session_preview = []
        self.session_applied = []

    def export_session_log(self):
        """Exporta el log de la sesión actual (preview + applied) a un archivo CSV temporal."""
        # exportar preview / applied como CSV temporal
        if not (self.session_preview or self.session_applied):
            messagebox.showinfo("Nada", "No hay entradas en la sesión para exportar.")
            return
        default = f"session_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=default, filetypes=[("CSV files","*.csv")])
        if not path:
            return
        import csv
        keys = ['timestamp','original_path','action','new_path','note']
        combined = [r for r in (self.session_preview + self.session_applied) if r.get('action') != 'VALIDADO_OK']
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            for r in combined:
                writer.writerow({k: r.get(k,'') for k in keys})
        messagebox.showinfo("Exportado", f"Log de sesión exportado a {path}")

    def log_to_session(self, original, action, newpath='', note=''):
        """Registra una entrada en la sesión actual y la muestra en el panel.
        Args:
            original (str): _ruta original del archivo_
            action (str): _acción realizada_
            newpath (str, optional): _nueva ruta si aplica_. Defaults to ''.
            note (str, optional): _nota descriptiva_. Defaults to ''.
        Returns:
            dict: _entrada de log creada_
        """
        entry = {
            'timestamp': datetime.datetime.now().isoformat(sep=' ', timespec='seconds'),
            'original_path': original,
            'action': action,
            'new_path': newpath,
            'note': note
        }
        # Crear texto acortado para mostrar en la tabla
        display_original = shorten_path(original, show_parts=3)
        display_new = shorten_path(newpath, show_parts=3) if newpath else ''
        # Insertar en treeview (mostrando ramas cortas)
        self.tree.insert('', 'end', values=(entry['timestamp'], display_original, entry['action'], display_new, entry['note']))
        return entry

    def run_scan(self, mode='preview'):
        """Ejecuta el escaneo y las operaciones según el modo seleccionado.
        Args:        
            mode: 'preview' -> no cambios en disco, no escribir log persistente
                'apply'   -> aplicar cambios y actualizar archivo log (si está establecido)
        """
        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Carpeta inválida", "Selecciona una carpeta raíz válida.")
            return

        pattern = self.pattern_var.get().strip()
        prefix_choice = self.prefix_var.get().strip()
        area_map_text = self.map_text.get('1.0', 'end').strip()
        area_map = parse_mapping_text(area_map_text)
        log_path = self.log_path_var.get().strip()

        if mode == 'apply' and not log_path:
            if not messagebox.askyesno("Confirmar sin log", "No seleccionaste un archivo de log. ¿Continuar sin persistir log?"):
                return

        preview_only = (mode == 'preview')
        if mode == 'apply':
            if not messagebox.askyesno("Confirmar aplicar cambios", "Se aplicarán cambios sobre archivos y se actualizará el log (si está seleccionado). ¿Continuar?"):
                return

        # limpiar panel de resultados de la ejecución previa
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.session_preview = []
        self.session_applied = []

        pattern = self.pattern_var.get().strip()
        pattern_re = pattern_to_regex(pattern)
        pattern_has_prefix = '{PREFIX}' in pattern
        pattern_has_parent = '{PARENT_LETTER}' in pattern

        selected_dirs = [os.path.normcase(os.path.abspath(k)) for k, v in self.folder_checks.items() if v.get()]
        any_selected = len(selected_dirs) > 0
        
        # recorrer
        for root_dir, dirs, files in os.walk(folder):
            # normalizar current root
            root_norm = os.path.normcase(os.path.abspath(root_dir))

            # Si hay subcarpetas seleccionadas, sólo procesar si root_dir está dentro de alguna seleccionada.
            if any_selected:
                ok = False
                for sd in selected_dirs:
                    # coincidencia exacta o subruta: sd == root_norm or root_norm startswith sd + sep
                    if root_norm == sd or root_norm.startswith(sd + os.sep):
                        ok = True
                        break
                if not ok:
                    # saltar este root_dir por completo
                    continue
            
            for fname in files:
                fullpath = os.path.join(root_dir, fname)
                try:
                    # 1) eliminar temporales
                    if self.delete_vars.get() and is_temp_file(fname):
                        if preview_only:
                            e = self.log_to_session(fullpath, 'DETECTADO_ELIMINAR_PREVIEW', '', 'archivo temporal detectado')
                            self.session_preview.append(e)
                        else:
                            try:
                                os.remove(fullpath)
                                e = self.log_to_session(fullpath, 'ELIMINADO', '', 'archivo temporal eliminado')
                                self.session_applied.append(e)
                            except Exception as ex:
                                e = self.log_to_session(fullpath, 'ERROR_ELIMINAR', '', str(ex))
                                if preview_only:
                                    self.session_preview.append(e)
                                else:
                                    self.session_applied.append(e)
                        continue

                    # 2) limpiar nombre (parche: no forzar lowercase y respetar pattern)
                    new_name_after_clean = fname
                    if self.clean_names_var.get():
                        # intentamos aprovechar pattern_re: si el archivo ya cumple el patrón,
                        # limpiamos solo el grupo {NAME}; si no coincide, limpiamos sólo la parte 'name' sin extensión
                        m_clean = None
                        try:
                            m_clean = pattern_re.fullmatch(fname)
                        except Exception:
                            m_clean = None

                        if m_clean:
                            # limpia solo el grupo NAME, sin forzar minusculas
                            groups_clean = m_clean.groupdict()
                            original_name_part = groups_clean.get('NAME', '')
                            processed_name_part, moved_number = process_name_for_standard(original_name_part, move_leading_number=True)
                            # si hay número movido, agregarlo al final del NAME
                            if moved_number:
                                processed_name_part = f"{processed_name_part}{'_' if not processed_name_part.endswith('_') else ''}{moved_number}"
                            processed_name_part = capitalize_first_word_keep_underscores(processed_name_part)
                            cleaned_name_part = processed_name_part
                            if cleaned_name_part != original_name_part:
                                # reconstruir usando los grupos originales (no alteramos prefix/area/parent/ext aquí)
                                # usamos la misma técnica de sustitución que se usa más adelante
                                newname = pattern
                                # si el pattern tiene placeholders, se sustituyen en su lugar
                                # algunos placeholders pueden no existir en pattern, pero replace no afecta si no están
                                newname = newname.replace('{PREFIX}', groups_clean.get('PREFIX',''))
                                newname = newname.replace('{AREA}', groups_clean.get('AREA',''))
                                newname = newname.replace('{NAME}', cleaned_name_part)
                                newname = newname.replace('{EXT}', groups_clean.get('EXT',''))
                                newname = newname.replace('{PARENT_LETTER}', groups_clean.get('PARENT_LETTER',''))
                                newname = re.sub(r'\.+', '.', newname).strip('. ')
                                new_full = os.path.join(root_dir, newname)
                                if preview_only:
                                    e = self.log_to_session(fullpath, 'RENOMBRAR_LIMPIEZA_ESTANDAR_PREVIEW', new_full, 'limpieza NAME en archivo que ya cumple patrón')
                                    self.session_preview.append(e)
                                else:
                                    target = safe_unique_path(root_dir, newname)
                                    try:
                                        os.rename(fullpath, target)
                                        e = self.log_to_session(fullpath, 'RENOMBRADO_LIMPIEZA_ESTANDAR', target, 'limpieza NAME en archivo que ya cumple patrón')
                                        self.session_applied.append(e)
                                        fullpath = target
                                        fname = os.path.basename(fullpath)
                                    except Exception as ex:
                                        e = self.log_to_session(fullpath, 'ERROR_RENOMBRAR', '', str(ex))
                                        if preview_only:
                                            self.session_preview.append(e)
                                        else:
                                            self.session_applied.append(e)
                        else:
                            # No coincide con patrón: limpiar solo la parte "name" (sin la extensión)
                            name_part, ext = os.path.splitext(fname)
                            cleaned = ascii_clean(name_part, lower=False)  # importante: no forzar lowercase
                            if cleaned != name_part:
                                new_name_after_clean = f"{cleaned}{ext}"
                                new_full = os.path.join(root_dir, new_name_after_clean)
                                if preview_only:
                                    e = self.log_to_session(fullpath, 'RENOMBRAR_LIMPIEZA_PREVIEW', new_full, 'limpieza de caracteres (no cambia prefijos/areas)')
                                    self.session_preview.append(e)
                                else:
                                    target = safe_unique_path(root_dir, new_name_after_clean)
                                    try:
                                        os.rename(fullpath, target)
                                        e = self.log_to_session(fullpath, 'RENOMBRADO_LIMPIEZA', target, 'limpieza de caracteres (no cambia prefijos/areas)')
                                        self.session_applied.append(e)
                                        fullpath = target
                                        fname = os.path.basename(fullpath)
                                    except Exception as ex:
                                        e = self.log_to_session(fullpath, 'ERROR_RENOMBRAR', '', str(ex))
                                        if preview_only:
                                            self.session_preview.append(e)
                                        else:
                                            self.session_applied.append(e) 

                    # 3) aplicar estándar
                    m = pattern_re.fullmatch(fname)
                    if m:
                        groups = m.groupdict()
                        original_prefix = groups.get('PREFIX', '')
                        original_area = groups.get('AREA', '')
                        original_name_part = groups.get('NAME', '')
                        original_ext = groups.get('EXT', '')
                        original_parent_letter = groups.get('PARENT_LETTER', '')

                        # --- parche: si el PREFIX contiene un número al inicio (ej "11 R"), extraerlo y pasarlo al NAME ----------
                        # detectar secuencia numérica al inicio del prefix: "123 R", "24 y 25 R", "8-21R", "24,25-R"
                        moved_number_from_prefix = None
                        mp = re.match(r'^\s*([0-9]+(?:\s*(?:[\-–\.,]|y|and|&)\s*[0-9]+)*)\s*[-_\.\s:]*(.*)$',
                                    original_prefix, flags=re.IGNORECASE)
                        if mp:
                            raw_num_seq = mp.group(1)
                            rest_pref = mp.group(2).strip()
                            normalized = re.sub(r'[\s]*?(?:[\-–\.,]|y|and|&)[\s]*?', '_', raw_num_seq, flags=re.IGNORECASE)
                            normalized = re.sub(r'_+', '_', normalized).strip('_ ')
                            if normalized:
                                moved_number_from_prefix = normalized
                            original_prefix = rest_pref or original_prefix  # si no queda texto, mantenemos original (evita vacío)
                        
                        # Procesar NAME una única vez (extrae número si viene al inicio del NAME)
                        processed_name_part, moved_number_from_name = process_name_for_standard(original_name_part, move_leading_number=True)

                        # Combinar números detectados (name primero, prefix segundo) y añadirlos sufijo al NAME
                        combined_number_parts = []
                        if moved_number_from_name:
                            combined_number_parts.append(moved_number_from_name)
                        if moved_number_from_prefix:
                            combined_number_parts.append(moved_number_from_prefix)
                        if combined_number_parts:
                            suffix_number = '_'.join(combined_number_parts)
                            if processed_name_part:
                                processed_name_part = f"{processed_name_part}{'_' if not processed_name_part.endswith('_') else ''}{suffix_number}"
                            else:
                                processed_name_part = suffix_number

                        cleaned_name_part = processed_name_part  # lo que usaremos en reconstrucción

                        # obtener la abreviacion esperada en esta ruta (según el mapeo)
                        expected_area = find_area_abbr_in_path(root_dir, area_map) or ''
                        # Solo usar expected_prefix si el pattern lo requiere
                        expected_prefix = prefix_choice if pattern_has_prefix else original_prefix

                        # obtener la letra de la carpeta inmediata (esperada) sólo si pattern lo requiere
                        parent_folder = os.path.basename(root_dir) or ''
                        expected_parent_letter = (parent_folder[0].upper() if parent_folder else '') if pattern_has_parent else original_parent_letter

                        # decidir si debemos cambiar PREFIX/AREA/PARENT_LETTER (o sólo NAME)
                        needs_prefix_change = pattern_has_prefix and (expected_prefix and original_prefix != expected_prefix)
                        needs_area_change = (expected_area and original_area != expected_area)
                        needs_parent_change = pattern_has_parent and (expected_parent_letter and original_parent_letter != expected_parent_letter)
                        needs_name_change = (cleaned_name_part != original_name_part)

                        if needs_prefix_change or needs_area_change or needs_name_change or needs_parent_change:
                            # definir variables antes de la reconstrucción (si toca cambiar)
                            prefix_to_use = expected_prefix if (pattern_has_prefix and expected_prefix) else original_prefix
                            area_to_use = expected_area if expected_area else original_area
                            parent_to_use = expected_parent_letter if (pattern_has_parent and expected_parent_letter) else original_parent_letter
                            ext_to_use = original_ext

                            # reconstruir según el pattern (sustituir placeholders)
                            newname = pattern
                            if pattern_has_prefix:
                                newname = newname.replace('{PREFIX}', prefix_to_use)
                            newname = newname.replace('{AREA}', area_to_use)
                            newname = newname.replace('{NAME}', cleaned_name_part)
                            newname = newname.replace('{EXT}', ext_to_use)
                            if pattern_has_parent:
                                newname = newname.replace('{PARENT_LETTER}', parent_to_use)
                            newname = re.sub(r'\.+', '.', newname).strip('. ')
                            new_full = os.path.join(root_dir, newname)

                            # log / acciones
                            action_preview = 'ESTANDARIZAR_REPARAR_PREFIJO_AREA_PREVIEW' if preview_only else 'ESTANDARIZAR_REPARAR_PREFIJO_AREA'
                            if pattern_has_parent:
                                action_preview = action_preview.replace('PREFIJO_AREA', 'PREFIJO_AREA_PARENT')

                            note_parts = []
                            if needs_prefix_change:
                                note_parts.append(f'prefix: {groups.get("PREFIX","")} -> {prefix_to_use}')
                            if needs_area_change:
                                note_parts.append(f'area: {original_area} -> {area_to_use}')
                            if needs_parent_change:
                                note_parts.append(f'parent_letter: {original_parent_letter} -> {parent_to_use}')
                            if needs_name_change:
                                note_parts.append('limpieza NAME')
                            note = '; '.join(note_parts)

                            if preview_only:
                                e = self.log_to_session(fullpath, action_preview, new_full, note)
                                self.session_preview.append(e)
                            else:
                                target = safe_unique_path(root_dir, newname)
                                try:
                                    os.rename(fullpath, target)
                                    e = self.log_to_session(fullpath, action_preview.replace('_PREVIEW',''), target, note)
                                    self.session_applied.append(e)
                                    fullpath = target
                                    fname = os.path.basename(fullpath)
                                except Exception as ex:
                                    e = self.log_to_session(fullpath, 'ERROR_RENOMBRAR', '', str(ex))
                                    if preview_only:
                                        self.session_preview.append(e)
                                    else:
                                        self.session_applied.append(e)
                        else:
                            pass  # ya cumple estándar, no hacer nada
                    else:
                        # NO cumple el patrón: aplicar la lógica normal de construcción
                        newname, note = build_standard_name(fname, root_dir, pattern, prefix_choice, area_map)
                        if newname != fname:
                            new_full = os.path.join(root_dir, newname)
                            if preview_only:
                                e = self.log_to_session(fullpath, 'ESTANDARIZAR_PREVIEW', new_full, note)
                                self.session_preview.append(e)
                            else:
                                target = safe_unique_path(root_dir, newname)
                                try:
                                    os.rename(fullpath, target)
                                    e = self.log_to_session(fullpath, 'ESTANDARIZADO', target, note)
                                    self.session_applied.append(e)
                                except Exception as ex:
                                    e = self.log_to_session(fullpath, 'ERROR_ESTANDARIZAR', '', str(ex))
                                    if preview_only:
                                        self.session_preview.append(e)
                                    else:
                                        self.session_applied.append(e)
                        else:
                            # si build_standard_name devolvió el mismo nombre, no mostramos ni guardamos
                            pass
                except Exception as ex:
                    e = self.log_to_session(fullpath, 'ERROR_GENERAL', '', str(ex))
                    if preview_only:
                        self.session_preview.append(e)
                    else:
                        self.session_applied.append(e)

        # Si aplicamos, persistir session_applied al archivo de log si existe ruta
        if mode == 'apply' and self.session_applied and log_path:
            try:
                append_log_entries_to_file(self.session_applied, log_path)
                messagebox.showinfo("Listo", f"Operaciones aplicadas y log actualizado en:\n{log_path}")
            except Exception as ex:
                messagebox.showwarning("Aplicado pero error guardando log", f"Cambios aplicados, pero hubo un error al actualizar el log:\n{ex}")
        else:
            if mode == 'apply':
                messagebox.showinfo("Listo", "Operaciones aplicadas (no se actualizó log persistente).")
            else:
                messagebox.showinfo("Vista previa generada", "Se generó la vista previa en el panel (no se modificó nada).")

    def clear_folder_checkboxes(self):
        """Remueve todos los checkbuttons y variables previas."""
        for widget in self.subdirs_inner.winfo_children():
            widget.destroy()
        self.folder_checks.clear()

    def populate_folder_checkboxes(self, root_folder: str):
        """Lista subcarpetas de primer nivel y crea checkbuttons para cada una."""
        self.clear_folder_checkboxes()
        try:
            # listar solo carpetas de primer nivel
            with os.scandir(root_folder) as it:
                dirs = [entry for entry in it if entry.is_dir()]
        except Exception:
            dirs = []

        # si no hay subcarpetas, mostrar la del mismo root (opcional)
        if not dirs:
            # mostrar la propia carpeta raíz como opción
            var = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(self.subdirs_inner, text=os.path.basename(root_folder) or root_folder, variable=var)
            cb.pack(anchor='w', pady=2, padx=2)
            self.folder_checks[os.path.abspath(root_folder)] = var
            return

        # crear checkbuttons ordenadas por nombre
        for d in sorted(dirs, key=lambda x: x.name.lower()):
            path_abs = os.path.abspath(d.path)
            var = tk.BooleanVar(value=True)  # por defecto: seleccionadas
            label = f"{d.name}"
            cb = ttk.Checkbutton(self.subdirs_inner, text=label, variable=var)
            cb.pack(anchor='w', pady=1, padx=2)
            self.folder_checks[path_abs] = var

    def select_all_subdirs(self, select=True):
        """Seleccionar o deseleccionar todas las subcarpetas en la lista."""
        for v in self.folder_checks.values():
            v.set(select)

    def invert_subdirs(self):
        """Invertir selección."""
        for v in self.folder_checks.values():
            v.set(not v.get())
# ---------------------------
# Ejecutar
# ---------------------------
def main():
    root = tk.Tk()
    root.geometry('1200x750')
    app = FileCleanerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()